
from __future__ import annotations

from io import BytesIO
from math import ceil
from pathlib import Path
from typing import Any, Dict, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class CalculationError(Exception):
    pass


TAX_SYSTEM_OPTIONS = {
    "ОСНО": 22.0,
    "УСН 6%": 6.0,
    "УСН доходы-расходы 15%": 15.0,
}

PRODUCT_COLUMN_MAP = {
    "sku": "SKU",
    "артикул": "SKU",
    "название": "Product Name",
    "название товара": "Product Name",
    "наименование": "Product Name",
    "товар": "Product Name",
    "product_name": "Product Name",
    "product name": "Product Name",
    "cost_rub": "Cost RUB",
    "себестоимость": "Cost RUB",
    "себестоимость (руб)": "Cost RUB",
    "себестоимость (₽)": "Cost RUB",
    "себестоимость, ₽": "Cost RUB",
    "себестоимость руб": "Cost RUB",
    "weight_kg": "Weight kg",
    "вес, кг": "Weight kg",
    "вес кг": "Weight kg",
    "вес": "Weight kg",
    "вес (кг)": "Weight kg",
    "length_cm": "Length cm",
    "длина, см": "Length cm",
    "длина см": "Length cm",
    "длина": "Length cm",
    "длина (см)": "Length cm",
    "width_cm": "Width cm",
    "ширина, см": "Width cm",
    "ширина см": "Width cm",
    "ширина": "Width cm",
    "ширина (см)": "Width cm",
    "height_cm": "Height cm",
    "высота, см": "Height cm",
    "высота см": "Height cm",
    "высота": "Height cm",
    "высота (см)": "Height cm",
    "model": "Model",
    "модель": "Model",
    "manual_category": "Manual Category",
    "категория": "Manual Category",
    "notes": "Notes",
    "комментарий": "Notes",
}

REQUIRED_PRODUCT_COLUMNS = ["SKU", "Product Name", "Cost RUB", "Weight kg", "Length cm", "Width cm", "Height cm"]

DEFAULT_SETTINGS = {
    "default_model": "FBS",
    "tax_system": "УСН 6%",
    "target_margin_pct": 20.0,
    "buyout_rate_pct": 95.0,
    "ad_rate_pct": 8.0,
    "defect_rate_pct": 1.0,
    "transfer_schedule": "Еженедельно, с отсрочкой в 2 недели",
    "reverse_same_locality_share_pct": 0.0,
    "fbs_avg_orders_per_day": 20,
    "avg_items_per_order": 1.2,
    "fby_storage_days": 30,
    "fby_storage_cost_rub_per_day": 0.0,
    "fby_inbound_self_delivery_rub_per_unit": 0.0,
    "default_commission_rate_pct": 8.0,
    "kgt_weight_threshold_kg": 25.0,
    "kgt_sum_sides_threshold_cm": 150.0,
}

RESULT_COLUMNS_RU = [
    "Артикул",
    "Название товара",
    "Модель",
    "Категория",
    "Источник категории",
    "Комиссия, %",
    "Себестоимость, ₽",
    "Вес, кг",
    "Длина, см",
    "Ширина, см",
    "Высота, см",
    "Объем, л",
    "Объемный вес, кг",
    "Сумма сторон, см",
    "КГТ",
    "Целевая маржинальность, %",
    "Налоговая система",
    "Налог, %",
    "Выкупаемость, %",
    "Реклама, %",
    "Брак / потери, %",
    "Ожидаемая выручка, ₽",
    "Комиссия ЯМ, ₽",
    "Перевод платежей, ₽",
    "Прием платежей, ₽",
    "Доставка покупателю, ₽",
    "Средняя миля, ₽",
    "Обратная логистика, ₽",
    "Обработка возвратов, ₽",
    "FBS: забор, ₽",
    "FBS: обработка, ₽",
    "FBY: хранение, ₽",
    "FBY: доставка до склада ЯМ, ₽",
    "Ожидаемая себестоимость проданных товаров, ₽",
    "Итого затрат, ₽",
    "Прибыль, ₽",
    "Маржа, % к выручке",
    "Рекомендованная цена, ₽",
    "Комментарий",
]


def _normalize_name(name: str) -> str:
    return str(name).strip().lower().replace("\n", " ").replace("  ", " ")


def _blankish(x: Any) -> bool:
    if x is None:
        return True
    try:
        if pd.isna(x):
            return True
    except Exception:
        pass
    return str(x).strip() == ""


def _clean_string(x: Any) -> str:
    if _blankish(x):
        return ""
    return str(x).strip()


def _to_float(x: Any, default: float = 0.0) -> float:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return default
    if isinstance(x, bool):
        return float(x)
    try:
        return float(str(x).replace(",", ".").strip())
    except Exception:
        return default


def _is_trueish(x: Any) -> bool:
    if isinstance(x, bool):
        return x
    return str(x).strip().lower() in {"1", "true", "yes", "y", "да", "x"}


def read_products_excel(file_like) -> pd.DataFrame:
    df = pd.read_excel(file_like, sheet_name=0)
    if df.empty:
        raise CalculationError("В файле товаров нет строк.")
    rename_map = {}
    for col in df.columns:
        rename_map[col] = PRODUCT_COLUMN_MAP.get(_normalize_name(col), col)
    df = df.rename(columns=rename_map)
    missing = [c for c in REQUIRED_PRODUCT_COLUMNS if c not in df.columns]
    if missing:
        missing_ru = {
            "SKU": "Артикул",
            "Product Name": "Название",
            "Cost RUB": "Себестоимость, ₽",
            "Weight kg": "Вес, кг",
            "Length cm": "Длина, см",
            "Width cm": "Ширина, см",
            "Height cm": "Высота, см",
        }
        raise CalculationError(
            "В файле товаров не хватает обязательных столбцов: " + ", ".join(missing_ru.get(c, c) for c in missing)
        )
    for optional in ["Model", "Manual Category", "Notes"]:
        if optional not in df.columns:
            df[optional] = None
    return df


def load_settings(file_or_path=None) -> Dict[str, Any]:
    params = DEFAULT_SETTINGS.copy()
    if file_or_path is None:
        return params
    wb = load_workbook(file_or_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        value = row[1]
        if key in params and value is not None:
            params[key] = value
    return params


def load_commissions(path: str | Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Commissions")
    df.columns = [str(c).strip() for c in df.columns]
    needed = {"Category", "Commission Rate %"}
    if not needed.issubset(df.columns):
        raise CalculationError("В файле commissions.xlsx нужен лист Commissions со столбцами Category и Commission Rate %.")
    df["Category_norm"] = df["Category"].astype(str).str.strip().str.lower()
    return df


def load_keyword_rules(path: str | Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Keyword Rules")
    df.columns = [str(c).strip() for c in df.columns]
    needed = {"Keyword", "Category", "Priority"}
    if not needed.issubset(df.columns):
        raise CalculationError("В файле keyword_rules.xlsx нужен лист Keyword Rules со столбцами Keyword, Category и Priority.")
    df["keyword_norm"] = df["Keyword"].astype(str).str.strip().str.lower()
    df = df.sort_values(["Priority", "Keyword"])
    return df


def load_tariffs(path: str | Path) -> Dict[str, Any]:
    general = pd.read_excel(path, sheet_name="General")
    payment_transfer = pd.read_excel(path, sheet_name="Payment Transfer")
    fbs_pickup = pd.read_excel(path, sheet_name="FBS Pickup")
    general_map = {str(k).strip(): v for k, v in zip(general["parameter"], general["value"])}
    payment_map = {
        str(row["frequency_label"]).strip(): float(row["rate_pct"])
        for _, row in payment_transfer.iterrows()
        if pd.notna(row["frequency_label"])
    }
    return {
        "general": general_map,
        "payment_transfer": payment_map,
        "fbs_pickup": fbs_pickup.to_dict("records"),
    }


def detect_category(product_name: str, keyword_rules_df: pd.DataFrame) -> Tuple[str, str]:
    name_norm = str(product_name).strip().lower()
    for _, row in keyword_rules_df.iterrows():
        kw = str(row["keyword_norm"]).strip()
        if kw and kw in name_norm:
            return row["Category"], f"по ключевому слову: {row['Keyword']}"
    return "UNMAPPED", "не определена"


def lookup_commission_rate(category: str, commissions_df: pd.DataFrame, default_rate_pct: float) -> Tuple[float, bool]:
    norm = str(category).strip().lower()
    match = commissions_df.loc[commissions_df["Category_norm"] == norm]
    if match.empty:
        return float(default_rate_pct), False
    val = match.iloc[0]["Commission Rate %"]
    if pd.isna(val):
        return float(default_rate_pct), False
    return float(val), True


def calc_volume_liters(length_cm: float, width_cm: float, height_cm: float) -> int:
    return max(1, ceil((float(length_cm) * float(width_cm) * float(height_cm)) / 1000.0))


def calc_volumetric_weight_kg(length_cm: float, width_cm: float, height_cm: float) -> float:
    return (float(length_cm) * float(width_cm) * float(height_cm)) / 5000.0


def calc_middle_mile_rub(volume_liters: int, tariffs: Dict[str, Any]) -> float:
    g = tariffs["general"]
    base = float(g["middle_mile_base_rub"])
    extra_1_160 = float(g["middle_mile_up_to_160_rub_per_l"])
    extra_160_plus = float(g["middle_mile_over_160_rub_per_l"])
    liters = int(volume_liters)
    if liters <= 1:
        return base
    if liters <= 160:
        return base + (liters - 1) * extra_1_160
    return base + 159 * extra_1_160 + (liters - 160) * extra_160_plus


def calc_fbs_pickup_fee_per_unit(avg_orders_per_day: float, avg_items_per_order: float, tariffs: Dict[str, Any]) -> float:
    day_orders = max(1.0, float(avg_orders_per_day))
    items_per_order = max(1.0, float(avg_items_per_order))
    chosen_fee = 0.0
    for row in tariffs["fbs_pickup"]:
        low = float(row["min_orders"])
        high = float(row["max_orders"]) if pd.notna(row["max_orders"]) else float("inf")
        if low <= day_orders <= high:
            chosen_fee = float(row["fee_rub"])
            break
    return chosen_fee / day_orders / items_per_order


def calc_fbs_processing_fee_per_unit(is_kgt: bool, volumetric_weight_kg: float, avg_items_per_order: float, tariffs: Dict[str, Any]) -> float:
    g = tariffs["general"]
    if is_kgt:
        return max(0.0, volumetric_weight_kg) * float(g["fbs_kgt_processing_rub_per_kg"])
    return float(g["fbs_regular_processing_rub_per_order"]) / max(1.0, float(avg_items_per_order))


def price_based_delivery_fee(price_rub: float, tariffs: Dict[str, Any]) -> float:
    g = tariffs["general"]
    rate = float(g["delivery_to_buyer_rate_pct"]) / 100.0
    cap = float(g["delivery_to_buyer_cap_rub"])
    return min(float(price_rub) * rate, cap)


def compute_profit_for_price(price_rub: float, product: Dict[str, Any], settings: Dict[str, Any], tariffs: Dict[str, Any], commission_rate_pct: float) -> Dict[str, float]:
    buyout = max(0.01, min(1.0, _to_float(settings["buyout_rate_pct"]) / 100.0))
    ad_rate = _to_float(settings["ad_rate_pct"]) / 100.0
    tax_rate = _to_float(TAX_SYSTEM_OPTIONS[settings["tax_system"]]) / 100.0
    defect_rate = _to_float(settings["defect_rate_pct"]) / 100.0
    transfer_rate = float(tariffs["payment_transfer"][settings["transfer_schedule"]]) / 100.0
    reverse_same_locality_share = _to_float(settings["reverse_same_locality_share_pct"]) / 100.0
    payment_acceptance = float(tariffs["general"]["payment_acceptance_rub_per_sku"])
    returns_handling = float(tariffs["general"]["returns_handling_rub_per_shipment"])

    price_rub = float(price_rub)
    cost_rub = _to_float(product["Cost RUB"])
    volume_liters = int(product["Volume liters"])
    middle_mile = calc_middle_mile_rub(volume_liters, tariffs)
    delivery_to_buyer = price_based_delivery_fee(price_rub, tariffs)

    model = product["Resolved Model"]
    fbs_pickup = 0.0
    fbs_processing = 0.0
    if model == "FBS":
        fbs_pickup = calc_fbs_pickup_fee_per_unit(settings["fbs_avg_orders_per_day"], settings["avg_items_per_order"], tariffs)
        fbs_processing = calc_fbs_processing_fee_per_unit(bool(product["Is KGT Resolved"]), _to_float(product["Volumetric weight kg"]), settings["avg_items_per_order"], tariffs)

    storage_cost = 0.0
    inbound_cost = 0.0
    if model == "FBY":
        storage_cost = _to_float(settings["fby_storage_days"]) * _to_float(settings["fby_storage_cost_rub_per_day"])
        inbound_cost = _to_float(settings["fby_inbound_self_delivery_rub_per_unit"])

    expected_revenue = price_rub * buyout
    commission = price_rub * (commission_rate_pct / 100.0) * buyout
    payment_transfer = price_rub * transfer_rate * buyout
    tax = price_rub * tax_rate * buyout
    ad_cost = price_rub * ad_rate * buyout
    defect_cost = cost_rub * defect_rate

    reverse_middle_mile = middle_mile * (1.0 - buyout) * (1.0 - reverse_same_locality_share)
    reverse_handling_cost = returns_handling * (1.0 - buyout)
    expected_cogs = cost_rub * buyout

    total_cost = (
        expected_cogs
        + commission
        + payment_transfer
        + payment_acceptance
        + delivery_to_buyer
        + middle_mile
        + reverse_middle_mile
        + reverse_handling_cost
        + tax
        + ad_cost
        + defect_cost
        + fbs_pickup
        + fbs_processing
        + storage_cost
        + inbound_cost
    )
    profit = expected_revenue - total_cost
    margin_pct = (profit / expected_revenue * 100.0) if expected_revenue > 0 else None
    return {
        "Expected Revenue RUB": expected_revenue,
        "Commission RUB": commission,
        "Payment Transfer RUB": payment_transfer,
        "Payment Acceptance RUB": payment_acceptance,
        "Delivery to Buyer RUB": delivery_to_buyer,
        "Middle Mile RUB": middle_mile,
        "Reverse Handling RUB": reverse_handling_cost,
        "Reverse Middle Mile RUB": reverse_middle_mile,
        "Tax RUB": tax,
        "Ads RUB": ad_cost,
        "Defect Cost RUB": defect_cost,
        "FBS Pickup RUB": fbs_pickup,
        "FBS Processing RUB": fbs_processing,
        "Storage RUB": storage_cost,
        "Inbound to Yandex WH RUB": inbound_cost,
        "Expected COGS RUB": expected_cogs,
        "Total Cost RUB": total_cost,
        "Profit RUB": profit,
        "Margin %": margin_pct,
        "Tax Rate %": tax_rate * 100.0,
    }


def solve_recommended_price(product: Dict[str, Any], settings: Dict[str, Any], tariffs: Dict[str, Any], commission_rate_pct: float, target_margin_pct: float, seed_price: float) -> float | None:
    target_margin = max(0.0, min(0.95, float(target_margin_pct) / 100.0))

    def margin_for(price: float) -> float:
        metrics = compute_profit_for_price(price, product, settings, tariffs, commission_rate_pct)
        revenue = metrics["Expected Revenue RUB"]
        if revenue <= 0:
            return -999.0
        return metrics["Profit RUB"] / revenue

    low = 1.0
    high = max(seed_price * 2.0 if seed_price and seed_price > 0 else 1000.0, 1000.0)
    tries = 0
    while margin_for(high) < target_margin and tries < 40:
        high *= 1.5
        tries += 1
        if high > 1_000_000:
            return None

    for _ in range(80):
        mid = (low + high) / 2.0
        if margin_for(mid) >= target_margin:
            high = mid
        else:
            low = mid
    return round(high, 2)


def calculate_unit_economics(products_df: pd.DataFrame, settings: Dict[str, Any], tariffs: Dict[str, Any], commissions_df: pd.DataFrame, keyword_rules_df: pd.DataFrame):
    results = []
    warnings = []

    for _, row in products_df.iterrows():
        product = row.to_dict()
        resolved_model = (_clean_string(product.get("Model")) or _clean_string(settings["default_model"])).upper()
        if resolved_model not in {"FBS", "FBY"}:
            resolved_model = settings["default_model"]

        manual_category = _clean_string(product.get("Manual Category"))
        if manual_category:
            category = manual_category
            category_source = "вручную"
        else:
            category, category_source = detect_category(product["Product Name"], keyword_rules_df)

        commission_rate_pct, commission_is_exact = lookup_commission_rate(category, commissions_df, float(settings["default_commission_rate_pct"]))

        sum_sides = _to_float(product["Length cm"]) + _to_float(product["Width cm"]) + _to_float(product["Height cm"])
        auto_kgt = _to_float(product["Weight kg"]) > _to_float(settings["kgt_weight_threshold_kg"]) or sum_sides > _to_float(settings["kgt_sum_sides_threshold_cm"])
        volume_liters = calc_volume_liters(product["Length cm"], product["Width cm"], product["Height cm"])
        volumetric_weight_kg = calc_volumetric_weight_kg(product["Length cm"], product["Width cm"], product["Height cm"])

        product["Resolved Model"] = resolved_model
        product["Resolved Category"] = category
        product["Category Source"] = category_source
        product["Commission Rate %"] = commission_rate_pct
        product["Commission Exact Match"] = commission_is_exact
        product["Volume liters"] = volume_liters
        product["Volumetric weight kg"] = round(volumetric_weight_kg, 2)
        product["Sum sides cm"] = round(sum_sides, 2)
        product["Is KGT Resolved"] = auto_kgt

        target_margin_pct = _to_float(settings["target_margin_pct"])
        rec_price = solve_recommended_price(
            product=product,
            settings=settings,
            tariffs=tariffs,
            commission_rate_pct=commission_rate_pct,
            target_margin_pct=target_margin_pct,
            seed_price=max(_to_float(product["Cost RUB"]) * 2.0, 500.0),
        )
        rec_metrics = compute_profit_for_price(rec_price, product, settings, tariffs, commission_rate_pct) if rec_price else {}

        results.append({
            "Артикул": product["SKU"],
            "Название товара": product["Product Name"],
            "Модель": resolved_model,
            "Категория": category,
            "Источник категории": category_source,
            "Комиссия, %": round(commission_rate_pct, 2),
            "Себестоимость, ₽": round(_to_float(product["Cost RUB"]), 2),
            "Вес, кг": round(_to_float(product["Weight kg"]), 3),
            "Длина, см": round(_to_float(product["Length cm"]), 2),
            "Ширина, см": round(_to_float(product["Width cm"]), 2),
            "Высота, см": round(_to_float(product["Height cm"]), 2),
            "Объем, л": volume_liters,
            "Объемный вес, кг": round(volumetric_weight_kg, 2),
            "Сумма сторон, см": round(sum_sides, 2),
            "КГТ": "Да" if auto_kgt else "Нет",
            "Целевая маржинальность, %": round(target_margin_pct, 2),
            "Налоговая система": settings["tax_system"],
            "Налог, %": round(rec_metrics.get("Tax Rate %", TAX_SYSTEM_OPTIONS[settings["tax_system"]]), 2),
            "Выкупаемость, %": round(_to_float(settings["buyout_rate_pct"]), 2),
            "Реклама, %": round(_to_float(settings["ad_rate_pct"]), 2),
            "Брак / потери, %": round(_to_float(settings["defect_rate_pct"]), 2),
            "Ожидаемая выручка, ₽": round(rec_metrics.get("Expected Revenue RUB", 0.0), 2),
            "Комиссия ЯМ, ₽": round(rec_metrics.get("Commission RUB", 0.0), 2),
            "Перевод платежей, ₽": round(rec_metrics.get("Payment Transfer RUB", 0.0), 2),
            "Прием платежей, ₽": round(rec_metrics.get("Payment Acceptance RUB", 0.0), 2),
            "Доставка покупателю, ₽": round(rec_metrics.get("Delivery to Buyer RUB", 0.0), 2),
            "Средняя миля, ₽": round(rec_metrics.get("Middle Mile RUB", 0.0), 2),
            "Обратная логистика, ₽": round(rec_metrics.get("Reverse Middle Mile RUB", 0.0), 2),
            "Обработка возвратов, ₽": round(rec_metrics.get("Reverse Handling RUB", 0.0), 2),
            "FBS: забор, ₽": round(rec_metrics.get("FBS Pickup RUB", 0.0), 2),
            "FBS: обработка, ₽": round(rec_metrics.get("FBS Processing RUB", 0.0), 2),
            "FBY: хранение, ₽": round(rec_metrics.get("Storage RUB", 0.0), 2),
            "FBY: доставка до склада ЯМ, ₽": round(rec_metrics.get("Inbound to Yandex WH RUB", 0.0), 2),
            "Ожидаемая себестоимость проданных товаров, ₽": round(rec_metrics.get("Expected COGS RUB", 0.0), 2),
            "Итого затрат, ₽": round(rec_metrics.get("Total Cost RUB", 0.0), 2),
            "Прибыль, ₽": round(rec_metrics.get("Profit RUB", 0.0), 2),
            "Маржа, % к выручке": round(rec_metrics.get("Margin %", 0.0), 2) if rec_metrics else None,
            "Рекомендованная цена, ₽": rec_price,
            "Комментарий": product.get("Notes"),
        })

        if category == "UNMAPPED":
            warnings.append({"Артикул": product["SKU"], "Что проверить": "Не удалось определить категорию по названию. Использована базовая комиссия."})
        elif not commission_is_exact:
            warnings.append({"Артикул": product["SKU"], "Что проверить": f"Для категории '{category}' не найдена точная комиссия. Использована базовая комиссия."})

    result_df = pd.DataFrame(results)
    warnings_df = pd.DataFrame(warnings)
    return result_df[RESULT_COLUMNS_RU], warnings_df


def _apply_sheet_style(ws, currency_cols=None, percent_cols=None, freeze="A2"):
    currency_cols = set(currency_cols or [])
    percent_cols = set(percent_cols or [])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[1].height = 32
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column in currency_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            if cell.column in percent_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'
    for idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col_cells:
            value = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 30)


def build_result_workbook(result_df: pd.DataFrame, warnings_df: pd.DataFrame, settings: Dict[str, Any], tariffs: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Результат"
    ws.append(list(result_df.columns))
    for row in result_df.itertuples(index=False):
        ws.append(list(row))

    settings_ws = wb.create_sheet("Параметры")
    settings_ws.append(["Параметр", "Значение"])
    settings_view = [
        ["Модель", settings["default_model"]],
        ["Налоговая система", settings["tax_system"]],
        ["Целевая маржинальность, %", settings["target_margin_pct"]],
        ["Выкупаемость, %", settings["buyout_rate_pct"]],
        ["Реклама, %", settings["ad_rate_pct"]],
        ["Брак / потери, %", settings["defect_rate_pct"]],
        ["Возвраты без обратной средней мили, %", settings["reverse_same_locality_share_pct"]],
        ["FBS: среднее число заказов в день", settings["fbs_avg_orders_per_day"]],
        ["Среднее число товаров в заказе", settings["avg_items_per_order"]],
        ["FBY: срок хранения, дней", settings["fby_storage_days"]],
        ["FBY: хранение, ₽ в день за 1 шт.", settings["fby_storage_cost_rub_per_day"]],
        ["FBY: наша доставка до склада ЯМ, ₽/шт.", settings["fby_inbound_self_delivery_rub_per_unit"]],
    ]
    for row in settings_view:
        settings_ws.append(row)

    warn_ws = wb.create_sheet("Что проверить")
    if warnings_df.empty:
        warn_ws.append(["Статус"])
        warn_ws.append(["Замечаний нет"])
    else:
        warn_ws.append(list(warnings_df.columns))
        for row in warnings_df.itertuples(index=False):
            warn_ws.append(list(row))

    source_ws = wb.create_sheet("Источники")
    source_ws.append(["Название", "URL"])
    for title, url in [
        ("Тарифы FBS", "https://yandex.ru/support/marketplace/ru/introduction/rates/models/fbs"),
        ("Тарифы FBY", "https://yandex.ru/support/marketplace/ru/introduction/rates/models/fby"),
        ("Прием и перевод платежей", "https://yandex.ru/support/marketplace/ru/introduction/rates/acquiring"),
        ("Калькулятор Яндекс Маркета", "https://partner.market.yandex.ru/welcome/cost-calculator"),
    ]:
        source_ws.append([title, url])

    currency_headers = {
        "Себестоимость, ₽", "Ожидаемая выручка, ₽", "Комиссия ЯМ, ₽", "Перевод платежей, ₽",
        "Прием платежей, ₽", "Доставка покупателю, ₽", "Средняя миля, ₽", "Обратная логистика, ₽",
        "Обработка возвратов, ₽", "FBS: забор, ₽", "FBS: обработка, ₽", "FBY: хранение, ₽",
        "FBY: доставка до склада ЯМ, ₽", "Ожидаемая себестоимость проданных товаров, ₽", "Итого затрат, ₽",
        "Прибыль, ₽", "Рекомендованная цена, ₽",
    }
    percent_headers = {"Комиссия, %", "Целевая маржинальность, %", "Налог, %", "Выкупаемость, %", "Реклама, %", "Брак / потери, %", "Маржа, % к выручке"}
    header_to_idx = {cell.value: cell.column for cell in ws[1]}
    currency_cols = {header_to_idx[h] for h in currency_headers if h in header_to_idx}
    percent_cols = {header_to_idx[h] for h in percent_headers if h in header_to_idx}
    _apply_sheet_style(ws, currency_cols=currency_cols, percent_cols=percent_cols)
    _apply_sheet_style(settings_ws)
    _apply_sheet_style(warn_ws)
    _apply_sheet_style(source_ws)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
